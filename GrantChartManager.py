"""
Project Plan Visualization Tool (Robust Version)
This script handles Vertex42 Gantt charts more reliably and provides better error messages.
"""
import argparse
import logging
import os
import sys
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)

class ProjectVisualizer:
    """Robust project visualizer for Vertex42 Gantt charts"""
    
    def __init__(self, file_path, sheet_name='Project schedule'):
        self.file_path = os.path.abspath(file_path)
        self.sheet_name = sheet_name
        self.df = None
        self.validated = False
        self.logger = logging.getLogger('ProjectVisualizer')
        self.logger.info(f"Initialized with file: {self.file_path}")
        
    def load_data(self):
        """Load project data with enhanced error handling"""
        try:
            # Verify file exists
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"File not found: {self.file_path}")
            
            self.logger.info(f"Loading data from {self.file_path} (sheet: {self.sheet_name})")
            
            # Load workbook
            wb = load_workbook(self.file_path, data_only=True)
            
            # Verify sheet exists
            if self.sheet_name not in wb.sheetnames:
                available_sheets = ", ".join(wb.sheetnames)
                raise ValueError(
                    f"Sheet '{self.sheet_name}' not found. "
                    f"Available sheets: {available_sheets}"
                )
            
            sheet = wb[self.sheet_name]
            self.logger.info(f"Successfully accessed sheet: {self.sheet_name}")
            
            # Find data starting row (Vertex42 templates start at row 6)
            START_ROW = 6
            data = []
            row_index = START_ROW
            
            while True:
                # Get task cell (column B)
                task_cell = f'B{row_index}'
                task_value = sheet[task_cell].value
                
                # Stop when we find an empty task cell
                if not task_value:
                    break
                
                # Get other values
                assigned_to = sheet[f'C{row_index}'].value
                progress = sheet[f'D{row_index}'].value
                start_date = sheet[f'E{row_index}'].value
                end_date = sheet[f'F{row_index}'].value
                
                # Skip section headers (rows without dates)
                if not start_date:
                    row_index += 1
                    continue
                
                data.append({
                    'Task': task_value,
                    'Assigned To': assigned_to,
                    'Progress': progress,
                    'Start': start_date,
                    'End': end_date
                })
                
                row_index += 1
            
            self.df = pd.DataFrame(data)
            self.logger.info(f"Loaded {len(self.df)} tasks")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading data: {str(e)}")
            # Provide troubleshooting tips
            self.logger.info("Troubleshooting tips:")
            self.logger.info("1. Verify the file path is correct")
            self.logger.info("2. Ensure the sheet name matches your project sheet")
            self.logger.info("3. Check that tasks start at row 6 with dates in columns E and F")
            return False
    
    def validate_data(self):
        """Validate and clean the project data"""
        if self.df is None or self.df.empty:
            self.logger.error("No data to validate")
            return False
        
        try:
            # Convert dates with error handling
            self.df['Start'] = pd.to_datetime(self.df['Start'], errors='coerce')
            self.df['End'] = pd.to_datetime(self.df['End'], errors='coerce')
            
            # Remove rows with invalid dates
            initial_count = len(self.df)
            self.df = self.df.dropna(subset=['Start', 'End'])
            removed_count = initial_count - len(self.df)
            if removed_count > 0:
                self.logger.warning(f"Removed {removed_count} tasks with invalid dates")
            
            # Handle missing values
            self.df['Assigned To'].fillna('Unassigned', inplace=True)
            self.df['Progress'].fillna(0, inplace=True)
            
            # Ensure progress is between 0-100%
            self.df['Progress'] = pd.to_numeric(self.df['Progress'], errors='coerce')
            self.df['Progress'] = self.df['Progress'].clip(0, 1)
            
            # Fix end dates that are before start dates
            invalid_dates = self.df[self.df['End'] < self.df['Start']]
            if not invalid_dates.empty:
                self.logger.warning(f"Fixed {len(invalid_dates)} tasks with end date before start date")
                self.df.loc[self.df['End'] < self.df['Start'], 'End'] = self.df['Start']
            
            # Calculate durations
            self.df['Duration'] = (self.df['End'] - self.df['Start']).dt.days + 1
            self.df['Completed End'] = self.df['Start'] + pd.to_timedelta(
                (self.df['Duration'] * self.df['Progress']).astype(int), 
                unit='d'
            )
            
            self.validated = True
            self.logger.info("Data validation complete")
            return True
            
        except Exception as e:
            self.logger.error(f"Validation error: {str(e)}")
            return False
    
    def create_gantt_chart(self, output_dir):
        """Create an interactive Gantt chart visualization"""
        if not self.validated:
            self.logger.error("Data not validated")
            return False
        
        try:
            self.logger.info("Creating Gantt chart visualization")
            
            # Create figure
            fig = go.Figure()
            
            # Add planned duration bars
            fig.add_trace(go.Bar(
                y=self.df['Task'],
                x=self.df['Duration'],
                base=self.df['Start'],
                orientation='h',
                name='Planned',
                marker=dict(color='rgba(100,100,100,0.3)'),
                hoverinfo='text',
                hovertext=[
                    f"<b>{row['Task']}</b><br>"
                    f"Owner: {row['Assigned To']}<br>"
                    f"Start: {row['Start'].strftime('%Y-%m-%d')}<br>"
                    f"End: {row['End'].strftime('%Y-%m-%d')}<br>"
                    f"Progress: {row['Progress']*100:.0f}%"
                    for _, row in self.df.iterrows()
                ]
            ))
            
            # Add completed progress bars
            completed_days = (self.df['Completed End'] - self.df['Start']).dt.days
            fig.add_trace(go.Bar(
                y=self.df['Task'],
                x=completed_days,
                base=self.df['Start'],
                orientation='h',
                name='Completed',
                marker=dict(color='rgba(0,150,0,0.7)'),
                hoverinfo='skip'
            ))
            
            # Add milestones (0-1 day tasks)
            milestones = self.df[self.df['Duration'] <= 1]
            if not milestones.empty:
                fig.add_trace(go.Scatter(
                    x=milestones['Start'] + timedelta(hours=12),
                    y=milestones['Task'],
                    mode='markers',
                    marker=dict(symbol='diamond', size=15, color='red'),
                    name='Milestones',
                    hoverinfo='text',
                    hovertext=[f"<b>{row['Task']}</b>" for _, row in milestones.iterrows()]
                ))
            
            # Update layout
            fig.update_layout(
                title='Project Gantt Chart',
                barmode='overlay',
                height=600 + len(self.df) * 20,
                xaxis_title='Timeline',
                yaxis_title='Tasks',
                yaxis=dict(autorange="reversed"),
                xaxis=dict(
                    type='date',
                    tickformat='%b %d',
                    rangeslider=dict(visible=True)
                ),
                legend=dict(
                    orientation='h', 
                    yanchor='bottom', 
                    y=1.02, 
                    xanchor='right', 
                    x=1
                ),
                hovermode='closest'
            )
            
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # Save outputs
            html_path = os.path.join(output_dir, 'gantt_chart.html')
            fig.write_html(html_path)
            self.logger.info(f"Saved interactive Gantt chart to {html_path}")
            
            png_path = os.path.join(output_dir, 'gantt_chart.png')
            fig.write_image(png_path, width=1200, height=800)
            self.logger.info(f"Saved Gantt chart image to {png_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating Gantt chart: {str(e)}")
            return False

def main():
    """Command-line interface for project visualization"""
    parser = argparse.ArgumentParser(description='Project Plan Visualization Tool')
    parser.add_argument('--load', required=True, help='Path to Excel project file')
    parser.add_argument('--sheet', default='Project schedule', help='Sheet name containing project data')
    parser.add_argument('--gantt', action='store_true', help='Generate Gantt chart')
    parser.add_argument('--output', default='output', help='Output directory for visualizations')
    
    args = parser.parse_args()
    
    # Initialize visualizer
    visualizer = ProjectVisualizer(args.load, args.sheet)
    
    # Load and validate data
    if not visualizer.load_data():
        sys.exit(1)
    
    if not visualizer.validate_data():
        sys.exit(1)
    
    # Generate requested visualizations
    if args.gantt:
        visualizer.create_gantt_chart(args.output)
    
    logging.info("Process completed successfully")

if __name__ == "__main__":
    main()
