import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt

# 1. Loop through all Excel files in the folder and extract data
def extract_and_consolidate(folder, sheet_name=None):
    all_data = []
    for fname in os.listdir(folder):
        if fname.endswith('.xlsx') and not fname.startswith('~$'):
            fpath = os.path.join(folder, fname)
            try:
                wb = load_workbook(fpath, data_only=True)
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                data = ws.values
                cols = next(data)
                df = pd.DataFrame(data, columns=cols)
                df['SourceFile'] = fname
                all_data.append(df)
            except Exception as e:
                print(f"Error reading {fname}: {e}")
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return pd.DataFrame()

# 2. Write consolidated data to master Excel file
def write_master_excel(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidated Data'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Add table
    tab = Table(displayName="ConsolidatedTable", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(out_path)

# 3. Create pivot table and chart in Excel
def add_pivot_and_chart(excel_path, pivot_col, value_col):
    wb = load_workbook(excel_path)
    ws = wb['Consolidated Data']
    # Create a summary sheet
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    summary = wb.create_sheet('Summary')
    # Simple pivot: sum by pivot_col
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df[1:]
    pivot = df.groupby(pivot_col)[value_col].sum().reset_index()
    for r in dataframe_to_rows(pivot, index=False, header=True):
        summary.append(r)
    # Add conditional formatting (top 10%)
    max_row = summary.max_row
    summary.conditional_formatting.add(f'B2:B{max_row}',
        CellIsRule(operator='greaterThan', formula=[f'LARGE(B2:B{max_row},ROUND(COUNT(B2:B{max_row})*0.1,0))'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
    # Add error highlighting (negative values)
    summary.conditional_formatting.add(f'B2:B{max_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')))
    wb.save(excel_path)

# 4. Create a chart image and insert into Excel
def add_chart_image(excel_path, pivot_col, value_col):
    wb = load_workbook(excel_path)
    ws = wb['Summary']
    data = list(ws.values)
    cols = data[0]
    df = pd.DataFrame(data[1:], columns=cols)
    plt.figure(figsize=(8,4))
    plt.bar(df[pivot_col], pd.to_numeric(df[value_col], errors='coerce'))
    plt.title(f'{value_col} by {pivot_col}')
    plt.ylabel(value_col)
    plt.xticks(rotation=45)
    img_path = excel_path.replace('.xlsx', '_chart.png')
    plt.tight_layout()
    plt.savefig(img_path)
    plt.close()
    return img_path

# 5. Save as PDF (requires Excel/Windows)
def save_as_pdf(excel_path, pdf_path):
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
    except Exception as e:
        print(f"PDF export failed: {e}")

# 6. Main automation function
def generate_dashboard(folder, sheet_name=None, pivot_col='SourceFile', value_col=None):
    df = extract_and_consolidate(folder, sheet_name)
    if df.empty:
        print("No data found.")
        return
    if not value_col:
        value_col = df.columns[-2]  # Guess a numeric column
    master_path = os.path.join(folder, 'MasterDashboard.xlsx')
    write_master_excel(df, master_path)
    add_pivot_and_chart(master_path, pivot_col, value_col)
    chart_img = add_chart_image(master_path, pivot_col, value_col)
    pdf_path = master_path.replace('.xlsx', '.pdf')
    save_as_pdf(master_path, pdf_path)
    print(f"Dashboard saved as {master_path} and {pdf_path}")
    print(f"Chart image: {chart_img}")

# Example usage:
# generate_dashboard('path_to_folder', sheet_name='Sheet1', pivot_col='SourceFile', value_col='Sales')
