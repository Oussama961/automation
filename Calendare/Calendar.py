"""
python calendar_tool.py [--load PATH] [--add-event DATE TITLE]
                       [--update-event DATE OLD_TITLE NEW_TITLE]
                       [--remove-event DATE TITLE]
                       [--batch DATES_FILE]
                       [--summary]
                       [--list-events]
                       [--output OUTPUT_PATH]
                       [--sheet-name SHEET_NAME]
                       [--verbose]
                       [--create-sample SAMPLE_PATH]


Load your existing calendar and add a single event
python calendar_tool.py --load calendar.xlsx \
                        --add-event 2025-07-01 "Project Kickoff"


Batch-add events from a CSV or text file
python calendar_tool.py --load calendar.xlsx \
                        --batch dates.txt \
                        --output updated_calendar.xlsx


Generate (or regenerate) the Event Summary sheet:
python calendar_tool.py --load calendar.xlsx \
                        --summary


Update or remove an event:
python calendar_tool.py --load calendar.xlsx \
                        --update-event 2025-07-01 "Project Kickoff" "Kickoff Meeting"
python calendar_tool.py --load calendar.xlsx \
                        --remove-event 2025-07-01 "Kickoff Meeting"


List all events in the terminal
python calendar_tool.py --load calendar.xlsx \
                        --list-events

Create a fresh sample calendar:
python calendar_tool.py --create-sample sample_calendar.xlsx


More detailed log output (for debugging), add --verbose (or -v):
python calendar_tool.py --load calendar.xlsx --add-event 2025-07-01 "Project Kickoff" --verbose



"""


import argparse
import csv
import logging
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any, Union
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('calendar_automation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class CalendarException(Exception):
    """Custom exception for calendar operations"""
    pass


class ExcelCalendarManager:
    def __init__(self, workbook_path: str, calendar_sheet_name: str = "Calendar"):
        self.workbook_path = Path(workbook_path)
        self.calendar_sheet_name = calendar_sheet_name
        self.workbook = None
        self.calendar_sheet = None
        self.events = {}  # Dictionary to store events {date: {cell_address: event_data}}
        
        self.event_styles = {
            'default': {
                'fill': PatternFill(start_color="FAB07F", end_color="FF6600", fill_type="solid"),
                'font': Font(bold=True, color="000000")
            },
            'important': {
                'fill': PatternFill(start_color="FA5252", end_color="FF0000", fill_type="solid"),
                'font': Font(bold=True, color="FFFFFF")
            },
            'meeting': {
                'fill': PatternFill(start_color="5CFF5C", end_color="00FF00", fill_type="solid"),
                'font': Font(bold=True, color="000000")
            }
        }
        
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """Load the Excel workbook and calendar sheet."""
        try:
            if not self.workbook_path.exists():
                raise CalendarException(f"Workbook not found: {self.workbook_path}")
            
            self.workbook = openpyxl.load_workbook(self.workbook_path)
            
            if self.calendar_sheet_name not in self.workbook.sheetnames:
                self.calendar_sheet = self.workbook.create_sheet(self.calendar_sheet_name)
                logger.info(f"Created new calendar sheet: {self.calendar_sheet_name}")
            else:
                self.calendar_sheet = self.workbook[self.calendar_sheet_name]            
            logger.info(f"Successfully loaded workbook: {self.workbook_path}")
            
        except Exception as e:
            raise CalendarException(f"Failed to load workbook: {str(e)}")
    
    def extract_dates(self) -> Dict[str, datetime]:
        date_cells = {}
        try:
            for row in self.calendar_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        date_obj = self._parse_date(cell.value)
                        if date_obj:
                            cell_address = f"{get_column_letter(cell.column)}{cell.row}"
                            date_cells[cell_address] = date_obj
                            logger.debug(f"Found date in {cell_address}: {date_obj}")
            
            logger.info(f"Extracted {len(date_cells)} date cells from calendar")
            return date_cells
            
        except Exception as e:
            logger.error(f"Error extracting dates: {str(e)}")
            return {}
    
    def _parse_date(self, value: Any) -> Optional[datetime]:
        """
        Parse various date formats into datetime objects.
        
        Args:
            value: Cell value to parse
            
        Returns:
            datetime object or None if parsing fails
        """
        if isinstance(value, datetime):
            return value
        elif isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        elif isinstance(value, str):
            # Try common date formats
            date_formats = [
                "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
            ]
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(value.strip(), fmt)
                except ValueError:
                    continue
        
        return None
    
    def add_event(self, target_date: Union[str, datetime], event_title: str, 
                  cell_address: Optional[str] = None, style: str = 'default') -> bool:
        """
        Add an event to a specific date.
        
        Args:
            target_date: Date as string (YYYY-MM-DD) or datetime object
            event_title: Title of the event
            cell_address: Specific cell address (optional)
            style: Style name for the event
            
        Returns:
            True if event was added successfully
        """
        try:
            # Parse target date
            if isinstance(target_date, str):
                date_obj = datetime.strptime(target_date, "%Y-%m-%d")
            else:
                date_obj = target_date
            
            # Find or create cell for the date
            if cell_address:
                target_cell = self.calendar_sheet[cell_address]
            else:
                target_cell = self._find_or_create_date_cell(date_obj)
            
            # Store event data
            date_key = date_obj.strftime("%Y-%m-%d")
            if date_key not in self.events:
                self.events[date_key] = {}
            
            cell_addr = f"{get_column_letter(target_cell.column)}{target_cell.row}"
            self.events[date_key][cell_addr] = {
                'title': event_title,
                'style': style,
                'datetime': date_obj
            }
            
            # Update cell content and apply styling
            if target_cell.value and isinstance(target_cell.value, str):
                target_cell.value = f"{target_cell.value}\n{event_title}"
            else:
                target_cell.value = f"{date_obj.strftime('%Y-%m-%d')}\n{event_title}"
            
            self._apply_event_style(target_cell, style)
            
            logger.info(f"Added event '{event_title}' to {date_key} at {cell_addr}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to add event: {str(e)}")
            return False
    
    def _find_or_create_date_cell(self, date_obj: datetime) -> openpyxl.cell.Cell:
        """Find existing date cell or create a new one."""
        # First, try to find existing date cell
        date_cells = self.extract_dates()
        for cell_addr, cell_date in date_cells.items():
            if cell_date.date() == date_obj.date():
                return self.calendar_sheet[cell_addr]
        
        # Create new cell in next available row
        next_row = self.calendar_sheet.max_row + 1
        target_cell = self.calendar_sheet.cell(row=next_row, column=1)
        target_cell.value = date_obj.strftime("%Y-%m-%d")
        
        return target_cell
    
    def _apply_event_style(self, cell: openpyxl.cell.Cell, style_name: str) -> None:
        """Apply styling to an event cell."""
        style = self.event_styles.get(style_name, self.event_styles['default'])
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def update_event(self, target_date: str, old_title: str, new_title: str) -> bool:
        """
        Update an existing event.
        
        Args:
            target_date: Date string (YYYY-MM-DD)
            old_title: Current event title
            new_title: New event title
            
        Returns:
            True if event was updated successfully
        """
        try:
            if target_date not in self.events:
                logger.warning(f"No events found for date: {target_date}")
                return False
            
            for cell_addr, event_data in self.events[target_date].items():
                if event_data['title'] == old_title:
                    event_data['title'] = new_title
                    
                    # Update cell content
                    cell = self.calendar_sheet[cell_addr]
                    if cell.value:
                        cell.value = cell.value.replace(old_title, new_title)
                    
                    logger.info(f"Updated event from '{old_title}' to '{new_title}' on {target_date}")
                    return True
            
            logger.warning(f"Event '{old_title}' not found on {target_date}")
            return False
            
        except Exception as e:
            logger.error(f"Failed to update event: {str(e)}")
            return False
    
    def remove_event(self, target_date: str, event_title: str) -> bool:
        """
        Remove an event from a specific date.
        
        Args:
            target_date: Date string (YYYY-MM-DD)
            event_title: Title of the event to remove
            
        Returns:
            True if event was removed successfully
        """
        try:
            if target_date not in self.events:
                logger.warning(f"No events found for date: {target_date}")
                return False
            
            for cell_addr, event_data in list(self.events[target_date].items()):
                if event_data['title'] == event_title:
                    # Remove from events dictionary
                    del self.events[target_date][cell_addr]
                    
                    # Update cell content
                    cell = self.calendar_sheet[cell_addr]
                    if cell.value:
                        cell.value = cell.value.replace(f"\n{event_title}", "")
                        cell.value = cell.value.replace(event_title, "")
                    
                    # Reset styling if no more events
                    if not self.events[target_date]:
                        cell.fill = PatternFill()
                        cell.font = Font()
                        del self.events[target_date]
                    
                    logger.info(f"Removed event '{event_title}' from {target_date}")
                    return True
            
            logger.warning(f"Event '{event_title}' not found on {target_date}")
            return False
            
        except Exception as e:
            logger.error(f"Failed to remove event: {str(e)}")
            return False
    
    def batch_add_events(self, dates_file: str, default_event_title: str = "Event") -> int:
        """
        Add events from a batch file (CSV or text).
        
        Args:
            dates_file: Path to file containing dates
            default_event_title: Default title for events
            
        Returns:
            Number of events successfully added
        """
        added_count = 0
        
        try:
            dates_path = Path(dates_file)
            if not dates_path.exists():
                raise CalendarException(f"Dates file not found: {dates_file}")
            
            dates_list = self._parse_dates_file(dates_path)
            
            for date_str in dates_list:
                try:
                    date_obj = datetime.strptime(date_str.strip(), "%Y-%m-%d")
                    if self.add_event(date_obj, default_event_title):
                        added_count += 1
                except ValueError:
                    logger.warning(f"Invalid date format: {date_str}")
                    continue
            
            logger.info(f"Batch added {added_count} events from {dates_file}")
            return added_count
            
        except Exception as e:
            logger.error(f"Failed to batch add events: {str(e)}")
            return added_count
    
    def _parse_dates_file(self, file_path: Path) -> List[str]:
        """Parse dates from CSV or text file."""
        dates = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                if file_path.suffix.lower() == '.csv':
                    csv_reader = csv.reader(f)
                    for row in csv_reader:
                        if row:  # Skip empty rows
                            dates.append(row[0])  # Assume dates are in first column
                else:
                    # Treat as text file with one date per line
                    dates = [line.strip() for line in f if line.strip()]
            
            return dates
            
        except Exception as e:
            logger.error(f"Error parsing dates file: {str(e)}")
            return []
    
    def generate_summary_sheet(self) -> None:
        try:
            # Create or get summary sheet
            summary_sheet_name = "Event Summary"
            if summary_sheet_name in self.workbook.sheetnames:
                summary_sheet = self.workbook[summary_sheet_name]
                # Clear existing content
                summary_sheet.delete_rows(1, summary_sheet.max_row)
            else:
                summary_sheet = self.workbook.create_sheet(summary_sheet_name)
            
            # Headers
            headers = ["Date", "Event Title", "Cell Address", "Link to Calendar"]
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Sort events by date and populate summary
            all_events = []
            for date_str, events_dict in self.events.items():
                for cell_addr, event_data in events_dict.items():
                    all_events.append({
                        'date': date_str,
                        'title': event_data['title'],
                        'cell_address': cell_addr,
                        'datetime': event_data['datetime']
                    })
            
            # Sort by datetime
            all_events.sort(key=lambda x: x['datetime'])
            
            # Populate summary sheet
            for row, event in enumerate(all_events, 2):
                summary_sheet.cell(row=row, column=1).value = event['date']
                summary_sheet.cell(row=row, column=2).value = event['title']
                summary_sheet.cell(row=row, column=3).value = event['cell_address']
                
                # Create hyperlink to calendar cell
                link_cell = summary_sheet.cell(row=row, column=4)
                link_cell.value = "Go to Calendar"
                link_cell.hyperlink = f"#{self.calendar_sheet_name}!{event['cell_address']}"
                link_cell.font = Font(color="0000FF", underline="single")
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Generated summary sheet with {len(all_events)} events")
            
        except Exception as e:
            logger.error(f"Failed to generate summary sheet: {str(e)}")
    
    def save_workbook(self, output_path: Optional[str] = None) -> bool:
        """
        Save the workbook to file.
        
        Args:
            output_path: Optional output path, defaults to original path
            
        Returns:
            True if saved successfully
        """
        try:
            save_path = Path(output_path) if output_path else self.workbook_path
            self.workbook.save(save_path)
            logger.info(f"Saved workbook to: {save_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to save workbook: {str(e)}")
            return False
    
    def get_events_summary(self) -> Dict[str, List[str]]:
        """Get a summary of all events organized by date."""
        summary = {}
        for date_str, events_dict in self.events.items():
            summary[date_str] = [event_data['title'] for event_data in events_dict.values()]
        return summary


def create_sample_calendar(file_path: str) -> None:
    """Create a sample calendar file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendar"
    
    # Add sample dates
    sample_dates = [
        "2024-01-15", "2024-02-20", "2024-03-10",
        "2024-04-05", "2024-05-12", "2024-06-18"
    ]
    
    for i, date_str in enumerate(sample_dates, 1):
        ws.cell(row=i, column=1).value = date_str
    
    wb.save(file_path)
    logger.info(f"Created sample calendar: {file_path}")


def parse_arguments() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Excel Calendar Automation Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
  # Load calendar and add an event
  python calendar.py --load calendar.xlsx --add-event "2024-12-25" "Christmas"
  
  # Batch add events from CSV
  python calendar.py --load calendar.xlsx --batch dates.csv --output new_calendar.xlsx
  
  # Create sample calendar
  python calendar.py --create-sample sample_calendar.xlsx
        """
    )
    
    parser.add_argument('--load', type=str, help='Path to Excel calendar file')
    parser.add_argument('--add-event', nargs=2, metavar=('DATE', 'TITLE'),
                       help='Add event: DATE (YYYY-MM-DD) and TITLE')
    parser.add_argument('--update-event', nargs=3, metavar=('DATE', 'OLD_TITLE', 'NEW_TITLE'),
                       help='Update event: DATE, OLD_TITLE, NEW_TITLE')
    parser.add_argument('--remove-event', nargs=2, metavar=('DATE', 'TITLE'),
                       help='Remove event: DATE and TITLE')
    parser.add_argument('--batch', type=str, help='Batch add events from CSV/text file')
    parser.add_argument('--output', type=str, help='Output file path')
    parser.add_argument('--create-sample', type=str, help='Create sample calendar file')
    parser.add_argument('--summary', action='store_true', help='Generate event summary sheet')
    parser.add_argument('--list-events', action='store_true', help='List all events')
    parser.add_argument('--sheet-name', type=str, default='Calendar', help='Calendar sheet name')
    parser.add_argument('--verbose', '-v', action='store_true', help='Enable verbose logging')
    
    return parser.parse_args()


def main():
    """Main entry point for the calendar automation tool."""
    args = parse_arguments()
    
    # Set logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        # Handle sample calendar creation
        if args.create_sample:
            create_sample_calendar(args.create_sample)
            return 0
        
        # Validate required arguments
        if not args.load:
            print("Error: --load argument is required (unless using --create-sample)")
            return 1
        
        # Initialize calendar manager
        calendar_manager = ExcelCalendarManager(args.load, args.sheet_name)
        
        # Execute operations
        operations_performed = False
        
        if args.add_event:
            date_str, title = args.add_event
            if calendar_manager.add_event(date_str, title):
                print(f"✓ Added event '{title}' on {date_str}")
                operations_performed = True
        
        if args.update_event:
            date_str, old_title, new_title = args.update_event
            if calendar_manager.update_event(date_str, old_title, new_title):
                print(f"✓ Updated event on {date_str}")
                operations_performed = True
        
        if args.remove_event:
            date_str, title = args.remove_event
            if calendar_manager.remove_event(date_str, title):
                print(f"✓ Removed event '{title}' from {date_str}")
                operations_performed = True
        
        if args.batch:
            count = calendar_manager.batch_add_events(args.batch)
            print(f"✓ Batch added {count} events")
            operations_performed = True
        
        if args.summary:
            calendar_manager.generate_summary_sheet()
            print("✓ Generated event summary sheet")
            operations_performed = True
        
        if args.list_events:
            events = calendar_manager.get_events_summary()
            if events:
                print("\nEvents Summary:")
                for date_str, titles in sorted(events.items()):
                    print(f"  {date_str}: {', '.join(titles)}")
            else:
                print("No events found in calendar")
        
        # Save workbook if operations were performed
        if operations_performed:
            if calendar_manager.save_workbook(args.output):
                output_path = args.output or args.load
                print(f"✓ Saved calendar to: {output_path}")
        
        return 0
        
    except CalendarException as e:
        logger.error(f"Calendar error: {str(e)}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
