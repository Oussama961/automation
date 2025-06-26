import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt

def extract_and_consolidate_calendars(folder, sheet_name=None):
    all_data = []
    for fname in os.listdir(folder):
        if fname.endswith('.xlsx') and not fname.startswith('~$'):
            fpath = os.path.join(folder, fname)
            try:
                wb = load_workbook(fpath, data_only=True)
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                data = ws.values
                cols = next(data)
                df = pd.DataFrame(data, columns=cols)
                df['SourceFile'] = fname
                all_data.append(df)
            except Exception as e:
                print(f"Error reading {fname}: {e}")
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return pd.DataFrame()

def write_master_calendar_excel(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidated Calendar'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    tab = Table(displayName="CalendarTable", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(out_path)

def add_calendar_pivot_and_chart(excel_path, pivot_col, value_col):
    wb = load_workbook(excel_path)
    ws = wb['Consolidated Calendar']
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    summary = wb.create_sheet('Summary')
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df[1:]
    pivot = df.groupby(pivot_col)[value_col].count().reset_index()
    for r in dataframe_to_rows(pivot, index=False, header=True):
        summary.append(r)
    max_row = summary.max_row
    summary.conditional_formatting.add(f'B2:B{max_row}',
        CellIsRule(operator='greaterThan', formula=[f'LARGE(B2:B{max_row},ROUND(COUNT(B2:B{max_row})*0.1,0))'], fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
    wb.save(excel_path)

def add_calendar_chart_image(excel_path, pivot_col, value_col):
    wb = load_workbook(excel_path)
    ws = wb['Summary']
    data = list(ws.values)
    cols = data[0]
    df = pd.DataFrame(data[1:], columns=cols)
    plt.figure(figsize=(8,4))
    plt.bar(df[pivot_col], pd.to_numeric(df[value_col], errors='coerce'))
    plt.title(f'{value_col} count by {pivot_col}')
    plt.ylabel(f'Count of {value_col}')
    plt.xticks(rotation=45)
    img_path = excel_path.replace('.xlsx', '_calendar_chart.png')
    plt.tight_layout()
    plt.savefig(img_path)
    plt.close()
    return img_path

def save_calendar_as_pdf(excel_path, pdf_path):
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
    except Exception as e:
        print(f"PDF export failed: {e}")

def generate_calendar_dashboard(folder, sheet_name=None, pivot_col='SourceFile', value_col=None):
    df = extract_and_consolidate_calendars(folder, sheet_name)
    if df.empty:
        print("No calendar data found.")
        return
    if not value_col:
        value_col = df.columns[1]  # Guess a main event/column
    master_path = os.path.join(folder, 'MasterCalendarDashboard.xlsx')
    write_master_calendar_excel(df, master_path)
    add_calendar_pivot_and_chart(master_path, pivot_col, value_col)
    chart_img = add_calendar_chart_image(master_path, pivot_col, value_col)
    pdf_path = master_path.replace('.xlsx', '.pdf')
    save_calendar_as_pdf(master_path, pdf_path)
    print(f"Calendar dashboard saved as {master_path} and {pdf_path}")
    print(f"Chart image: {chart_img}")

# Example usage:
# generate_calendar_dashboard('path_to_calendare_folder', sheet_name='Calendar', pivot_col='SourceFile', value_col='Event Title')
